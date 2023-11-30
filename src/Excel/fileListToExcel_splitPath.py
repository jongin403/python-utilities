import os
import openpyxl
import sys
from datetime import datetime

def add_row(ws, row_data, last_folder_structure_list, row_num):
    for i, folder_name in enumerate(row_data):
        if i < len(last_folder_structure_list) and folder_name == last_folder_structure_list[i]:
            cell = ws.cell(row=row_num, column=i+1, value="")
        else:
            cell = ws.cell(row=row_num, column=i+1, value=folder_name)
    row_num += 1
    return row_num

# 각 파일 및 폴더에 대한 정보를 워크시트에 추가하는 함수
def add_file_folder_info(ws, root, names, folder_path, row_num, last_folder_structure_list):
    folder_structure = os.path.relpath(root, folder_path)
    folder_structure_list = folder_structure.split(os.path.sep)

    for name in names:
        if folder_structure == '.':
            row_data = [name]
        else:
            row_data = folder_structure_list + [name]
        row_num = add_row(ws, row_data, last_folder_structure_list, row_num)
        # 파일 이름이 있는 셀에만 하이퍼링크 추가
        if os.path.isfile(os.path.join(root, name)):
            cell = ws.cell(row=row_num, column=len(row_data))
            cell.hyperlink = os.path.abspath(os.path.join(root, name))

    return row_num, folder_structure_list

# 폴더 정보를 워크시트에 추가
def add_folder_info(ws, root, folder_path, row_num, last_folder_structure_list):
    names = sorted(os.listdir(root))
    items = [name for name in names if os.path.isdir(os.path.join(root, name)) or os.path.isfile(os.path.join(root, name))]

    for name in items:
        path = os.path.join(root, name)
        if os.path.isfile(path) or not os.listdir(path):  # 파일이거나 비어있는 폴더인 경우에만 출력
            row_num, last_folder_structure_list = add_file_folder_info(ws, root, [name], folder_path, row_num, last_folder_structure_list)
        if os.path.isdir(path) and os.listdir(path):  # 폴더 내에 파일이나 하위 폴더가 있을 경우 하위 폴더를 탐색
            row_num, last_folder_structure_list = add_folder_info(ws, path, folder_path, row_num, last_folder_structure_list)

    return row_num, last_folder_structure_list

def create_excel_file(folder_path):
    parent_folder_path = os.path.dirname(folder_path)
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")  # 파일 이름에 타임스탬프 추가
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "File list to Excel"

    row_num, _ = add_folder_info(ws, folder_path, folder_path, 1, [])  # 폴더 정보를 워크시트에 추가

    excel_file_name = f"FileList_{timestamp}.xlsx"
    excel_file_path = os.path.join(parent_folder_path, excel_file_name)
    wb.save(excel_file_path)

    return excel_file_path

def open_file(file_path):
    system = sys.platform
    if system.startswith("win"):
        os.system(f'start "" "{file_path}"')
    elif system == "darwin":  # macOS
        os.system(f'open "{file_path}"')
    else:
        print("Windows와 macOS만 지원합니다.")

def main():
    print("특정 폴더 아래에 있는 파일과 폴더 리스트를 Excel로 출력해주는 프로그램입니다.")
    print("폴더의 절대 경로를 입력하세요: ")
    folder_path = input()

    if not os.path.exists(folder_path):
        print("지정한 폴더 경로가 존재하지 않습니다.")
        return

    excel_file_path = create_excel_file(folder_path)
    print(f"Excel 파일이 생성되었습니다: {excel_file_path}")

    open_file(excel_file_path)

if __name__ == "__main__":
    main()