import pandas as pd
import openpyxl

TARGET_EXCEL_PATH = 'target.xlsx'
COMPARE_SHEET_NAME1 = '공종별 내역서(변경전)'
COMPARE_SHEET_NAME2 = '공종별내역서(변경후)'
RESULT_SHEET_NAME = 'Sheet3' # 시트명
LARGE_CATEGORY_INDEX = 0 # 대분류가 위치하는 인덱스
TARGET_ID_LIST = [0, 1, 2] # 구분을 위한 인덱스 리스트
IDENTIFIER_NAME = 'identifier' # 식별자 이름 설정

# 엑셀 파일에서 특정 시트를 읽고 전처리를 수행하는 함수
def preprocess_excel_sheet(sheet_name):
    df = pd.read_excel(TARGET_EXCEL_PATH, sheet_name=sheet_name, header=None)
    df.columns = range(df.shape[1]) # 칼럼명 변경
    df.fillna('', inplace=True) # 빈칸 처리
    df[IDENTIFIER_NAME] = combine_columns(df, TARGET_ID_LIST) # 비교용 구분자 생성
    return df

# 데이터프레임의 각 행에서 target_ids에 해당하는 인덱스의 열 값을 문자열로 결합하여 반환
def combine_columns(df, target_column_index_list):
    combined_columns = []
    for index, row in df.iterrows():
        if all(row[idx] != '' for idx in target_column_index_list):
            combined_str = ''.join(str(row[idx]) for idx in target_column_index_list)
            combined_columns.append(combined_str)
        else:
            combined_columns.append('')
    return combined_columns

# df1, df2 에서 LARGE_CATEGORY_INDEX 에 해당하는 열의 값이 일치하는 행을 찾아 반환하며 df1, df2 순으로 중복없는 값을 반환
def get_large_categories(df1, df2, target_index):
    column_values1 = df1.iloc[:, target_index].tolist()
    column_values2 = df2.iloc[:, target_index].tolist()
    filtered_values2 = [value for value in column_values2 if value not in column_values1]
    distinct_values = []
    for value in column_values1 + filtered_values2:
        if value == '':
            continue
        if value in distinct_values:
            continue
        distinct_values.append(value)
    return distinct_values

def get_compare_result(df1, df2, large_categories):
    results = [] # 결과를 담을 리스트 초기화

    # large_categories 를 순회하면서 df1 과 df2 에서 LARGE_CATEGORY_INDEX 에 해당하는 열의 값이 일치하는 행을 찾기
    for category in large_categories:
        df1_rows = df1[df1[LARGE_CATEGORY_INDEX] == category]
        df2_rows = df2[df2[LARGE_CATEGORY_INDEX] == category]
        # df1_rows 를 순회하면서 IDENTIFIER_NAME 에 해당하는 열의 값이 일치하는 df2_rows 값이 있을 경우, 없을 경우 구분
        for index, row1 in df1_rows.iterrows():
            identifier = row1[IDENTIFIER_NAME]
            matching_rows = df2_rows[df2_rows[IDENTIFIER_NAME] == identifier]
            # 일치하는 값이 있을 경우
            if not matching_rows.empty:
                matching_row = matching_rows.iloc[0]
                if matching_row.equals(row1):
                    continue
                # matching_rows 와 row1 을 비교하여 다른 값이 있을 경우
                # 첫번째 column 에 "당초" 표시 후 그 다음 칼럼부터 row1 와 동일한 데이터로 row 추가"
                row1_values = row1.values.tolist()
                row1_values.insert(0, "당초")  # 첫 번째 위치에 "당초"를 삽입
                results.append(row1_values)
                # 첫번째 column 에 "변경" 표시 후 그 다음 칼럼부터 matching_rows 와 동일한 데이터로 row 추가"
                matching_row_values = matching_row.values.tolist()
                matching_row_values.insert(0, "변경")  # 각 행의 첫 번째 위치에 "변경" 삽입
                results.append(matching_row_values)
            else:
                # 첫번째 column 에 "제거" 표시 후 그 다음 칼럼부터 동일한 데이터로 row 추가
                row1_values = row1.values.tolist()
                row1_values.insert(0, "제거")  # 첫 번째 위치에 "제거"를 삽입
                results.append(row1_values)
        # df2_rows 를 순회하면서 IDENTIFIER_NAME 에 해당하는 열의 값이 일치하는 df1_rows 값이 없을 경우, 구분
        for index, row2 in df2_rows.iterrows():
            identifier = row2[IDENTIFIER_NAME]
            matching_rows = df1_rows[df1_rows[IDENTIFIER_NAME] == identifier]
            # 일치하는 값이 없을 경우
            if matching_rows.empty:
                # 첫번째 column 에 "추가" 표시 후 그 다음 칼럼부터 동일한 데이터로 row 추가
                row2_values = row2.values.tolist()
                row2_values.insert(0, "추가")  # 첫 번째 위치에 "추가"를 삽입
                results.append(row2_values)

    #show_pretty_results(results)
    return results

def show_pretty_results(results):
    for row in results:
        print(*row, sep='\t')

# 결과를 엑셀 파일의 시트에 추가하는 함수
def append_results_to_excel(results, target_excel_path, sheet_name):
    # 엑셀 파일 읽기
    wb = openpyxl.load_workbook(target_excel_path)
    # 시트 선택
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    wb.create_sheet(sheet_name)
    sheet = wb[sheet_name]
    # 결과를 시트에 추가
    for row in results:
        sheet.append(row)
    # 변경사항 저장
    wb.save(target_excel_path)

def main():
    # 엑셀 파일에서 각 시트 읽기
    df1 = preprocess_excel_sheet(COMPARE_SHEET_NAME1)
    df2 = preprocess_excel_sheet(COMPARE_SHEET_NAME2)

    # 대분야 리스트
    large_categories = get_large_categories(df1, df2, LARGE_CATEGORY_INDEX)

    # 비교 및 결과 추가 함수 호출
    difference = get_compare_result(df1, df2, large_categories)

    append_results_to_excel(difference, TARGET_EXCEL_PATH, RESULT_SHEET_NAME)

if __name__ == "__main__":
    main()