import os
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import platform
from datetime import datetime

def create_excel_file(folder_path):
    # 부모 폴더 경로 추출
    parent_folder_path = os.path.dirname(folder_path)

    # 파일 이름에 타임스탬프 추가
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")

    # 새로운 Excel 워크북 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "File list to Excel"

    # 열 헤더 설정
    headers = ["파일/폴더 이름", "타입", "경로"]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"] = header
        ws[f"{col_letter}1"].font = Font(bold=True)

    # 폴더 구조를 순회하며 Excel 시트에 정보 추가
    row_num = 2
    for root, dirs, files in os.walk(folder_path):
        for name in dirs + files:
            ws[f"A{row_num}"] = name
            ws[f"B{row_num}"] = "Folder" if os.path.isdir(os.path.join(root, name)) else "File"
            ws[f"C{row_num}"] = os.path.abspath(os.path.join(root, name))

            # 파일/폴더 이름에 하이퍼링크 추가
            hyperlink = os.path.abspath(os.path.join(root, name))
            ws[f"A{row_num}"].hyperlink = hyperlink

            row_num += 1

    # 부모 폴더에 타임스탬프가 추가된 Excel 파일 저장
    excel_file_name = f"FileList_{timestamp}.xlsx"
    excel_file_path = os.path.join(parent_folder_path, excel_file_name)
    wb.save(excel_file_path)

    return excel_file_path

def open_file(file_path):
    system = platform.system()
    if system == "Windows":
        os.system(f'start "" "{file_path}"')
    elif system == "Darwin":  # macOS
        os.system(f'open "{file_path}"')
    else:
        print("Windows 와 macOS 만 지원합니다.")

def main():
    print("특정 폴더 아래에 있는 파일과 폴더 리스트를 Excel 로 출력해주는 프로그램입니다.")
    folder_path = input("폴더의 절대 경로를 입력하세요: ")

    if not os.path.exists(folder_path):
        print("지정한 폴더 경로가 존재하지 않습니다.")
        return

    excel_file_path = create_excel_file(folder_path)
    print(f"Excel 파일이 생성되었습니다: {excel_file_path}")

    open_file(excel_file_path)

if __name__ == "__main__":
    main()